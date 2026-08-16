"""
清洗 LBNL Queued Up 原始数据(用户提供的官方xlsx，emp.lbl.gov本身被Cloudflare拦截无法直接下载)，
产出 F09（本报告最关键的对照变量：并网队列时长/完成率/容量）用的处理后数据。

依赖：openpyxl (pip install openpyxl)
输入：data/raw/lbnl/LBNL_Queued_Up_thru2025.xlsx
输出：
  data/processed/F09_duration_ir_to_cod.csv   中位/p25/p75并网时长(月)，按投产年份，2005-2025
  data/processed/F09_duration_stages.csv      分阶段时长对照(IR→IA / IA→COD / IR→COD)，用于F10
  data/processed/F09_completion_rate.csv      完成率，按申请年份cohort
  data/processed/F09_annual_capacity.csv      年度新增申请容量(GW)，2000-2025
"""

import csv
from pathlib import Path

import openpyxl

RAW_XLSX = Path(__file__).resolve().parents[2] / "data" / "raw" / "lbnl" / "LBNL_Queued_Up_thru2025.xlsx"
OUT_DIR = Path(__file__).resolve().parents[2] / "data" / "processed"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def extract_duration_table(ws, header_row, id_col_name="In-Service Year"):
    """提取形如 [Year, n, mean, p25, Median, p75] 的时长分布表"""
    rows = []
    header = None
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 7)]
        if vals[0] in ("In-Service Year", "Year of IA"):
            header = vals
            continue
        if header and isinstance(vals[0], int) and 1990 <= vals[0] <= 2030:
            rows.append(dict(zip(["year", "n", "mean", "p25", "median", "p75"], vals[: len(header)])))
    return rows


def main():
    wb = openpyxl.load_workbook(RAW_XLSX, data_only=True)

    # 1. IR -> COD 中位并网时长（F09核心指标）
    ws = wb["37. IR to COD - all"]
    ir_to_cod = extract_duration_table(ws, None)
    out_csv = OUT_DIR / "F09_duration_ir_to_cod.csv"
    with out_csv.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["year", "n", "mean", "p25", "median", "p75"])
        writer.writeheader()
        writer.writerows(ir_to_cod)
    print(f"写入 {out_csv} ({len(ir_to_cod)} 行) — IR to COD, 2005-2025")

    # 2. 分阶段对照：IR->IA, IA->COD, IR->COD（F10用，判断哪个阶段是真正的瓶颈）
    ir_to_ia = extract_duration_table(wb["29. IR to IA - all"], None)
    ia_to_cod = extract_duration_table(wb["34. IA to COD - all"], None)
    stage_rows = []
    for r in ir_to_ia:
        stage_rows.append({"year": r["year"], "stage": "IR_to_IA", "median_months": r["median"], "n": r["n"]})
    for r in ia_to_cod:
        stage_rows.append({"year": r["year"], "stage": "IA_to_COD", "median_months": r["median"], "n": r["n"]})
    for r in ir_to_cod:
        stage_rows.append({"year": r["year"], "stage": "IR_to_COD", "median_months": r["median"], "n": r["n"]})
    stage_rows.sort(key=lambda x: (x["stage"], x["year"]))
    out_csv2 = OUT_DIR / "F09_duration_stages.csv"
    with out_csv2.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["year", "stage", "median_months", "n"])
        writer.writeheader()
        writer.writerows(stage_rows)
    print(f"写入 {out_csv2} ({len(stage_rows)} 行) — 分阶段时长对照")

    # 3. 完成率：按申请年份cohort的项目状态计数(as of end of 2025)
    ws3 = wb["23. Completion Rate Trend"]
    completion_rows = []
    for r in range(1, ws3.max_row + 1):
        year = ws3.cell(row=r, column=1).value
        if not (isinstance(year, int) and 1995 <= year <= 2030):
            continue
        active = ws3.cell(row=r, column=2).value or 0
        operational = ws3.cell(row=r, column=3).value or 0
        withdrawn = ws3.cell(row=r, column=4).value or 0
        suspended = ws3.cell(row=r, column=5).value or 0
        total = active + operational + withdrawn + suspended
        if total == 0:
            continue
        completion_rows.append(
            {
                "request_year": year,
                "count_active": active,
                "count_operational": operational,
                "count_withdrawn": withdrawn,
                "count_suspended": suspended,
                "total": total,
                "completion_rate_pct": round(operational / total * 100, 2),
                "withdrawal_rate_pct": round(withdrawn / total * 100, 2),
            }
        )
    out_csv3 = OUT_DIR / "F09_completion_rate.csv"
    with out_csv3.open("w", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "request_year", "count_active", "count_operational", "count_withdrawn",
                "count_suspended", "total", "completion_rate_pct", "withdrawal_rate_pct",
            ],
        )
        writer.writeheader()
        writer.writerows(completion_rows)
    print(f"写入 {out_csv3} ({len(completion_rows)} 行) — 完成率(按申请年份cohort，状态截至2025年底)")

    # 4. 年度新增申请容量(GW)，2000-2025
    ws4 = wb["05. Annual Requests"]
    capacity_rows = []
    for r in range(1, ws4.max_row + 1):
        year = ws4.cell(row=r, column=1).value
        if not (isinstance(year, int) and 1995 <= year <= 2030):
            continue
        n = ws4.cell(row=r, column=2).value
        cap_gw = ws4.cell(row=r, column=3).value
        capacity_rows.append({"request_year": year, "n_requests": n, "capacity_gw": round(cap_gw, 2)})
    out_csv4 = OUT_DIR / "F09_annual_capacity.csv"
    with out_csv4.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["request_year", "n_requests", "capacity_gw"])
        writer.writeheader()
        writer.writerows(capacity_rows)
    print(f"写入 {out_csv4} ({len(capacity_rows)} 行) — 年度新增申请容量")


if __name__ == "__main__":
    main()
