"""
把F02_business_ai_adoption的数据+图表导出成原生Excel工作簿（应用户要求，
下载到本地后图表本身可编辑，不是贴图）。

用法：python export_F02_xlsx.py
输出：figures/xlsx/F02_business_ai_adoption.xlsx
      Data工作表：原始数据+数据源说明；Chart工作表：原生Excel折线图
"""

import csv
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import Marker

SRC = Path(__file__).resolve().parents[2] / "data" / "processed" / "F02_business_ai_adoption.csv"
OUT_DIR = Path(__file__).resolve().parents[2] / "figures" / "xlsx"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT = OUT_DIR / "F02_business_ai_adoption.xlsx"

rows = list(csv.DictReader(open(SRC)))
by_date = defaultdict(dict)
for r in rows:
    by_date[r["date"]][r["metric"]] = float(r["pct"])

dates = sorted(by_date.keys())
metrics = [
    ("current_v1_goods_services", "当前使用(生产商品/服务场景)"),
    ("expected_v1_goods_services", "预期6个月后使用(同一问法)"),
    ("current_v2_any_function", "当前使用(任意业务环节,新问法)"),
    ("expected_v2_any_function", "预期6个月后使用(新问法)"),
]

wb = Workbook()

# ---- Data sheet ----
ws = wb.active
ws.title = "Data"
FONT_NAME = "Arial"

header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ["日期(双周采集期结束日)"] + [label for _, label in metrics]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[1].height = 32

for i, d in enumerate(dates, start=2):
    dt = datetime.strptime(d, "%Y-%m-%d")
    dc = ws.cell(row=i, column=1, value=dt)
    dc.number_format = "yyyy-mm-dd"
    dc.font = Font(name=FONT_NAME, size=10)
    dc.border = border
    for j, (mkey, _) in enumerate(metrics, start=2):
        val = by_date[d].get(mkey)
        cell = ws.cell(row=i, column=j, value=val)
        cell.font = Font(name=FONT_NAME, size=10)
        cell.number_format = "0.0"
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

ws.column_dimensions["A"].width = 18
for col in "BCDE":
    ws.column_dimensions[col].width = 24
ws.freeze_panes = "B2"

last_row = len(dates) + 1

# ---- Source/notes block ----
note_row = last_row + 3
notes = [
    "数据源：U.S. Census Bureau, Business Trends and Outlook Survey (BTOS)，核心AI问题，全国口径(州/行业/规模分层汇总后)",
    "API: https://www.census.gov/hfp/btos/api/periods/{period_id}/data (未在标准data.census.gov目录中，从BTOS官网前端JS逆向找到)",
    "访问日期：2026-08-16　覆盖：Period 31-106 (2023-09-11 至 2026-08-09)，73/76期成功(缺Period 85-87，问法切换窗口期)",
    "",
    "重要口径说明：",
    "2023-09-11至2025-11-16(Period 31-87)问法："
    "\"In the last two weeks, did this business use Artificial Intelligence (AI) in producing goods or services?\"",
    "2025-11-17起(Period 88+)问法："
    "\"In the last two weeks, did this business use Artificial Intelligence (AI) in any of its business functions?\"",
    "新问法范围从\"生产商品/服务环节\"扩大到\"企业任意业务环节\"(含营销、财务、HR等后台职能)，"
    "两者不是同一指标，图表和分析中不可连成一条线解读为连续趋势。",
    "",
    "验证：Period 31(2023-09-11~09-24) \"当前使用\"=3.7%，与NBER Working Paper No. 32319"
    "(Bonney et al., 2024, census.gov/hfp/btos/downloads/CES-WP-24-16.pdf之关联工作论文)摘要"
    "\"3.7% at the start of the collection in September 2023\"完全一致。",
]
for i, line in enumerate(notes):
    cell = ws.cell(row=note_row + i, column=1, value=line)
    cell.font = Font(name=FONT_NAME, size=9, italic=True, color="6B7280")
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
for i in range(len(notes)):
    ws.merge_cells(start_row=note_row + i, start_column=1, end_row=note_row + i, end_column=5)
    ws.row_dimensions[note_row + i].height = 15

# ---- Chart sheet ----
chart = LineChart()
chart.title = "企业AI采用率：\"在工作中使用AI\"的双周官方统计，2023–2026"
chart.style = 2
chart.y_axis.title = "%（受访企业中回答“是”的比例）"
chart.x_axis.title = "采集期结束日"
chart.x_axis.number_format = "yyyy-mm"
chart.x_axis.majorTimeUnit = "months"
chart.height = 12
chart.width = 26

colors = ["2A78D6", "86B6EF", "EB6834", "F3B08F"]  # blue solid, blue light, orange solid, orange light
dash = [None, "dash", None, "dash"]

cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
for idx, (mkey, label) in enumerate(metrics):
    col = idx + 2
    data_ref = Reference(ws, min_col=col, min_row=1, max_row=last_row)
    chart.add_data(data_ref, titles_from_data=True)

chart.set_categories(cats)

for idx, series in enumerate(chart.series):
    series.graphicalProperties.line.solidFill = colors[idx]
    series.graphicalProperties.line.width = 20000 if idx in (0, 2) else 14000
    if dash[idx]:
        series.graphicalProperties.line.dashStyle = dash[idx]
    series.marker = Marker(symbol="none")
    series.smooth = False

ws_chart = wb.create_sheet("Chart")
ws_chart.add_chart(chart, "B2")
ws_chart.sheet_view.showGridLines = False
title_cell = ws_chart.cell(row=1, column=1, value="见Data工作表获取原始数据与数据源说明")
title_cell.font = Font(name=FONT_NAME, size=9, italic=True, color="6B7280")

wb.save(OUT)
print(f"写入 {OUT}")
