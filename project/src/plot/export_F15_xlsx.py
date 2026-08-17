"""
把五大厂商资本开支——季度实际值(F15主图) + 年度前瞻(F15补充图，官方指引/分析师估计)
——导出成一个原生Excel工作簿(应用户要求，下载后图表可编辑，不是贴图)。

用法：python export_F15_xlsx.py
输出：figures/xlsx/F15_capex_actuals_and_outlook.xlsx
      "季度实际值"：2015Q1-2026Q2逐季资本开支(SEC EDGAR XBRL)
      "年度前瞻"：2025实际→2026官方指引→2027/2028分析师估计，含数据层级(tier)颜色区分
      "图表"：两张原生Excel图表(季度堆叠柱 + 年度前瞻分组柱)
"""

import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference

QUARTERLY_SRC = Path(__file__).resolve().parents[2] / "data" / "processed" / "F15_capex_quarterly.csv"
OUTLOOK_SRC = Path(__file__).resolve().parents[2] / "data" / "processed" / "F15_capex_outlook.csv"
OUT_DIR = Path(__file__).resolve().parents[2] / "figures" / "xlsx"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT = OUT_DIR / "F15_capex_actuals_and_outlook.xlsx"

FONT_NAME = "Arial"
COMPANIES = ["MSFT", "AMZN", "GOOGL", "META", "ORCL"]

TIER_FILL = {
    "actual": PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"),
    "official": PatternFill(start_color="DCEAFB", end_color="DCEAFB", fill_type="solid"),
    "analyst": PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid"),
    "none": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
}
TIER_LABEL = {
    "actual": "实际值",
    "official": "官方指引",
    "analyst": "分析师估计",
    "none": "无可靠数字",
}

header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row, headers, widths):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[row].height = 30
    for col, w in zip("ABCDEFGHIJKL", widths):
        ws.column_dimensions[col].width = w


wb = Workbook()

# ==================== Sheet 1: 季度实际值 ====================
ws1 = wb.active
ws1.title = "季度实际值"

rows = list(csv.DictReader(QUARTERLY_SRC.open()))
by_q = defaultdict(dict)
for r in rows:
    by_q[r["calendar_quarter"]][r["company"]] = int(r["capex_usd"]) / 1e9
quarters = sorted(by_q.keys())

style_header(ws1, 1, ["季度"] + [f"{c} ($B)" for c in COMPANIES] + ["五家合计 ($B)"],
             [10, 12, 12, 12, 12, 12, 14])

for i, q in enumerate(quarters, start=2):
    qc = ws1.cell(row=i, column=1, value=q)
    qc.font = Font(name=FONT_NAME, size=10)
    qc.border = border
    qc.alignment = Alignment(horizontal="center")
    for j, co in enumerate(COMPANIES, start=2):
        val = by_q[q].get(co)
        cell = ws1.cell(row=i, column=j, value=val)
        cell.font = Font(name=FONT_NAME, size=10)
        cell.number_format = "$#,##0.0"
        cell.border = border
    total_cell = ws1.cell(row=i, column=7, value=f"=SUM(B{i}:F{i})")
    total_cell.font = Font(name=FONT_NAME, size=10, bold=True)
    total_cell.number_format = "$#,##0.0"
    total_cell.border = border

last_row1 = len(quarters) + 1
ws1.freeze_panes = "B2"

note_row = last_row1 + 3
notes1 = [
    "数据源：SEC EDGAR XBRL companyfacts API，标签PaymentsToAcquirePropertyPlantAndEquipment"
    "(AMZN 2017年后改用PaymentsToAcquireProductiveAssets)",
    "https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json　访问日期：2026-08-16",
    "覆盖：按财季结束日期换算为日历季度；MSFT财年6月结束、ORCL财年5月结束，其余三家为自然年，均已对齐为日历季",
    "ORCL数据从2016Q2起(更早期财报未披露可比口径的capex)，故2015Q1-2016Q1的ORCL列为空，不是缺失数据",
    "已用各公司10-K年度合计交叉验证：如MSFT FY2025四个季度加总=$645.51亿，与10-K年度数字完全一致",
]
for i, line in enumerate(notes1):
    cell = ws1.cell(row=note_row + i, column=1, value=line)
    cell.font = Font(name=FONT_NAME, size=9, italic=True, color="6B7280")
    ws1.merge_cells(start_row=note_row + i, start_column=1, end_row=note_row + i, end_column=7)

# ==================== Sheet 2: 年度前瞻 ====================
ws2 = wb.create_sheet("年度前瞻")

outlook_rows = list(csv.DictReader(OUTLOOK_SRC.open()))

headers2 = ["公司", "期间", "类型", "口径说明", "下限($B)", "上限($B)", "点值($B)",
            "数据层级", "数据来源", "来源链接", "备注"]
style_header(ws2, 1, headers2, [8, 10, 10, 26, 10, 10, 10, 12, 34, 10, 60])

for i, r in enumerate(outlook_rows, start=2):
    tier = r["tier"]
    fill = TIER_FILL[tier]
    vals = [
        r["company"], r["period_label"], TIER_LABEL[tier], r["fiscal_note"],
        float(r["low_usd_b"]) if r["low_usd_b"] else None,
        float(r["high_usd_b"]) if r["high_usd_b"] else None,
        float(r["point_usd_b"]) if r["point_usd_b"] else None,
        tier, r["source_name"], r["source_url"], r["note"],
    ]
    for j, v in enumerate(vals, start=1):
        cell = ws2.cell(row=i, column=j, value=v)
        cell.font = Font(name=FONT_NAME, size=9.5, italic=(tier == "none"), color="6B7280" if tier == "none" else "000000")
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top") if j in (4, 9, 10, 11) else Alignment(vertical="top")
        if j in (5, 6, 7):
            cell.number_format = "$#,##0.0"
            cell.alignment = Alignment(horizontal="center", vertical="top")
    ws2.row_dimensions[i].height = 42

last_row2 = len(outlook_rows) + 1
ws2.freeze_panes = "B2"

note_row2 = last_row2 + 3
notes2 = [
    "数据源：实际值=SEC EDGAR XBRL(与季度实际值工作表同一批数据，日历年求和；ORCL用官方财年口径55.7B以便与FY指引可比)",
    "官方指引=各公司2026年最新一次财报电话会/新闻稿(CNBC/CFO Dive/Intellectia等财经媒体转引)",
    "2027/2028分析师估计=Morgan Stanley(大摩)对五大云厂商用同一套方法论的建模估计(经techtimes等财经媒体转引，"
    "无法直接访问原始研报核实，单一机构来源，不代表市场普遍共识)",
    "MSFT/ORCL用财年(MSFT:7-6月，ORCL:6-5月)，其余三家用自然年，不可跨公司直接比较\"期间\"列",
    "MSFT 2026指引从190B下修到175B是租赁会计分类调整，非真实缩减；ORCL FY2027的92.5B是gross口径，"
    "官方另给net(扣除客户预付/自供硬件报销约20-25B)约70B",
    "大摩自己的估计在2026年内也被多次上修：META 2027从185.6B上修到225B(+21%)、AMZN 2027上修15%至308B、"
    "2028上修29%至318B",
]
for i, line in enumerate(notes2):
    cell = ws2.cell(row=note_row2 + i, column=1, value=line)
    cell.font = Font(name=FONT_NAME, size=9, italic=True, color="6B7280")
    ws2.merge_cells(start_row=note_row2 + i, start_column=1, end_row=note_row2 + i, end_column=11)

# ---- 透视小表(供图表用)：公司 × 通用时间槽 ----
pivot_start = note_row2 + len(notes2) + 3
ws2.cell(row=pivot_start, column=1, value="图表用透视表（各公司口径不同，此处统一成4个通用槽位，具体期间见上表）").font = \
    Font(name=FONT_NAME, size=9.5, bold=True, italic=True, color="374151")

pivot_headers = ["公司", "最近实际", "官方指引(下一年)", "预测+2年(官方或分析师)", "预测+3年(分析师)"]
style_header(ws2, pivot_start + 1, pivot_headers, [10, 14, 18, 22, 18])

by_co = defaultdict(dict)
for r in outlook_rows:
    by_co[r["company"]][r["period_label"]] = r

PIVOT_MAP = {
    "MSFT": ["2025", "2026", "2027E", "FY2028"],
    "AMZN": ["2025", "2026", "2027E", "2028E"],
    "GOOGL": ["2025", "2026", "2027E", "2028"],
    "META": ["2025", "2026", "2027E", "2028E"],
    "ORCL": ["FY2026", "FY2027", "FY2028", None],
}
for i, co in enumerate(COMPANIES, start=pivot_start + 2):
    ws2.cell(row=i, column=1, value=co).font = Font(name=FONT_NAME, size=10, bold=True)
    ws2.cell(row=i, column=1).border = border
    for j, plabel in enumerate(PIVOT_MAP[co], start=2):
        r = by_co[co].get(plabel) if plabel else None
        val = float(r["point_usd_b"]) if (r and r["point_usd_b"]) else None
        cell = ws2.cell(row=i, column=j, value=val)
        cell.number_format = "$#,##0.0"
        cell.font = Font(name=FONT_NAME, size=10)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

pivot_last_row = pivot_start + 1 + len(COMPANIES)

# ==================== Sheet 3: 图表 ====================
ws3 = wb.create_sheet("图表")
ws3.sheet_view.showGridLines = False

chart1 = BarChart()
chart1.type = "col"
chart1.grouping = "stacked"
chart1.overlap = 100
chart1.title = "五大厂商资本开支季度实际值，2015Q1–2026Q2"
chart1.y_axis.title = "资本开支 ($B)"
chart1.x_axis.title = "季度"
chart1.height = 11
chart1.width = 32

cats1 = Reference(ws1, min_col=1, min_row=2, max_row=last_row1)
data1 = Reference(ws1, min_col=2, max_col=6, min_row=1, max_row=last_row1)
chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(cats1)

company_colors = {"MSFT": "2A78D6", "AMZN": "EB6834", "GOOGL": "1BAF7A", "META": "EDA100", "ORCL": "E87BA4"}
for series, co in zip(chart1.series, COMPANIES):
    series.graphicalProperties.solidFill = company_colors[co]

ws3.add_chart(chart1, "A1")

chart2 = BarChart()
chart2.type = "col"
chart2.grouping = "clustered"
chart2.title = "五大厂商资本开支前瞻：最近实际 → 官方指引 → 分析师估计"
chart2.y_axis.title = "资本开支 ($B)"
chart2.x_axis.title = "公司"
chart2.height = 11
chart2.width = 22

cats2 = Reference(ws2, min_col=1, min_row=pivot_start + 2, max_row=pivot_last_row)
data2 = Reference(ws2, min_col=2, max_col=5, min_row=pivot_start + 1, max_row=pivot_last_row)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)

slot_colors = ["9CA3AF", "2A78D6", "EB6834", "1BAF7A"]
for series, color in zip(chart2.series, slot_colors):
    series.graphicalProperties.solidFill = color

ws3.add_chart(chart2, "A24")

note = ws3.cell(row=47, column=1, value="见\"季度实际值\"\"年度前瞻\"工作表获取原始数据与数据源说明")
note.font = Font(name=FONT_NAME, size=9, italic=True, color="6B7280")

wb.save(OUT)
print(f"写入 {OUT}")
